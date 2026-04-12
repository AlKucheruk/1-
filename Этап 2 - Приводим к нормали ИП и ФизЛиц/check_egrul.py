import pandas as pd
import requests
import time
import re
from tqdm import tqdm
from datetime import datetime
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ================= НАСТРОЙКИ =================
INPUT_FILE = 'Исходные_Данные.xlsx'
SHEET_NAME = 'Лист1'

# Имя выходного файла с датой
today = datetime.now().strftime("%y.%m.%d")
OUTPUT_FILE = f'Обработанные_Данные_{today}.xlsx'

COL_INN = 'ИНН'
# =============================================

def search_by_inn(inn: str) -> dict:
    """Поиск по ИНН через egrul.nalog.ru"""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                      '(KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36'
    }
    
    try:
        # Первый запрос — токен
        r = requests.post('https://egrul.nalog.ru/', 
                          data={'query': inn.strip()}, 
                          headers=headers, timeout=15)
        r.raise_for_status()
        data = r.json()
        
        if data.get('captchaRequired'):
            return {'error': 'Капча'}
        
        token = data.get('t')
        if not token:
            return {'error': 'Нет токена'}
        
        time.sleep(0.8)
        
        # Второй запрос — результат
        r2 = requests.get(f'https://egrul.nalog.ru/search-result/{token}', 
                          headers=headers, timeout=15)
        r2.raise_for_status()
        return r2.json()
        
    except Exception as e:
        return {'error': str(e)[:120]}


def parse_egr_result(result: dict, fallback_name: str = '') -> dict:
    """Улучшенный парсинг результата"""
    def normalize_text(value) -> str:
        """Нормализация строк ФНС (в т.ч. неразрывные пробелы)."""
        text = '' if value is None else str(value)
        return text.replace('\xa0', ' ').replace('&nbsp;', ' ').strip()

    def format_person_name(value: str) -> str:
        """Приводит ФИО к виду: Фамилия Имя Отчество."""
        text = normalize_text(value)
        if not text:
            return ''
        return ' '.join(part.capitalize() for part in text.lower().split())

    def extract_date_by_label(text: str, label: str) -> str:
        """Достает дату формата дд.мм.гггг после указанной метки."""
        if not text:
            return ''
        m = re.search(rf'{re.escape(label)}\s*:?\s*(\d{{2}}\.\d{{2}}\.\d{{4}})', text, flags=re.IGNORECASE)
        return m.group(1) if m else ''

    if 'error' in result:
        return {
            'Это_ИП': False,
            'Статус': 'Ошибка запроса',
            'ОГРНИП': '',
            'Дата_присвоения_ОГРНИП': '',
            'Дата_прекращения_деятельности': '',
            'Название_ЕГР': '',
            'ЕГРЮЛ_ЕГРИП_Данные': f"Ошибка: {result['error']}"
        }
    
    rows = result.get('rows', [])
    if not rows:
        fallback_person_name = format_person_name(fallback_name)
        name_egr = f"ФЛ {fallback_person_name}" if fallback_person_name else "ФЛ"
        return {
            'Это_ИП': False,
            'Статус': 'Физическое лицо (никогда не было ИП)',
            'ОГРНИП': '',
            'Дата_присвоения_ОГРНИП': '',
            'Дата_прекращения_деятельности': '',
            'Название_ЕГР': name_egr,
            'ЕГРЮЛ_ЕГРИП_Данные': 'Данные не найдены в ЕГР'
        }
    
    row = rows[0]
    
    # Основные поля (ключи могут немного варьироваться)
    full_name = normalize_text(row.get('n', row.get('name', '')))
    ogrnip = normalize_text(row.get('o', row.get('ogrnip', '')))
    inn = normalize_text(row.get('i', row.get('inn', '')))
    
    # Дата регистрации / присвоения ОГРНИП
    date_reg = normalize_text(
        row.get('r', row.get('d', row.get('rd', row.get('dateReg', row.get('ДатаРег', '')))))
    )
    
    # Дата прекращения деятельности
    date_end = normalize_text(
        row.get('e', row.get('ce', row.get('dateEnd', row.get('ДатаПрекр', ''))))
    )

    # Фолбэк: иногда даты приходят только в "сыром" тексте карточки
    row_text = ' '.join(normalize_text(v) for v in row.values())
    if not date_reg:
        date_reg = extract_date_by_label(row_text, 'Дата присвоения ОГРНИП')
    if not date_end:
        date_end = extract_date_by_label(row_text, 'Дата прекращения деятельности')
    
    # Определяем статус
    has_ogrnip = len(ogrnip) == 15 and ogrnip.isdigit()
    
    if has_ogrnip:
        is_ip = True
        if date_end and date_end != 'None' and date_end != '':
            status = 'Ликвидировано (бывшее ИП)'
        else:
            status = 'Действующее ИП'
    else:
        is_ip = False
        status = 'Физическое лицо (никогда не было ИП)'

    # Для физлиц, если ФНС не вернула ФИО, берем из входного файла
    if not full_name and not is_ip:
        full_name = normalize_text(fallback_name)

    formatted_name = format_person_name(full_name)
    
    # Формируем красивое название
    if is_ip:
        name_egr = f"ИП {formatted_name}" if formatted_name else "ИП"
    else:
        name_egr = f"ФЛ {formatted_name}" if formatted_name else "ФЛ"
    
    # Подробная информация
    details = f"Наименование: {formatted_name or full_name}\n"
    details += f"ИНН: {inn}\n"
    if has_ogrnip:
        details += f"ОГРНИП: {ogrnip}\n"
    if date_reg:
        details += f"Дата присвоения ОГРНИП: {date_reg}\n"
    if date_end:
        details += f"Дата прекращения деятельности: {date_end}\n"
    
    return {
        'Это_ИП': is_ip,
        'Статус': status,
        'ОГРНИП': ogrnip,
        'Дата_присвоения_ОГРНИП': date_reg,
        'Дата_прекращения_деятельности': date_end,
        'Название_ЕГР': name_egr,
        'ЕГРЮЛ_ЕГРИП_Данные': details.strip()
    }


# ================= ЗАПУСК =================
print("Загружаем файл...")
df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME)

# Добавляем новые колонки
new_cols = ['Это_ИП', 'Статус', 'ОГРНИП', 'Дата_присвоения_ОГРНИП',
            'Дата_прекращения_деятельности', 'Название_ЕГР', 'ЕГРЮЛ_ЕГРИП_Данные']

for col in new_cols:
    if col not in df.columns:
        df[col] = ''

print(f"Начинаем обработку {len(df)} контрагентов по ИНН...")

for idx, row in tqdm(df.iterrows(), total=len(df), desc="Проверка по ИНН"):
    inn = str(row.get(COL_INN, '')).strip()
    
    if not inn or inn in ['nan', 'None', '']:
        df.at[idx, 'Статус'] = 'Нет ИНН для проверки'
        continue
    
    # Очищаем ИНН от лишних символов
    inn_clean = ''.join(filter(str.isdigit, inn))
    
    try:
        result = search_by_inn(inn_clean)
        fallback_name = row.get('НаименованиеКонтрагента', '')
        parsed = parse_egr_result(result, fallback_name=fallback_name)
        
        for key, value in parsed.items():
            df.at[idx, key] = value
            
    except Exception as e:
        df.at[idx, 'Статус'] = f'Ошибка: {str(e)[:80]}'
    
    time.sleep(0.85)  # безопасная пауза

# Сохраняем результат и форматируем как таблицу Excel
with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name=SHEET_NAME)
    ws = writer.book[SHEET_NAME]

    # В именах таблиц Excel нельзя использовать дефис, поэтому используем tbl_IP_FL
    table_name = 'tbl_IP_FL'
    last_col_letter = get_column_letter(ws.max_column)
    table_ref = f"A1:{last_col_letter}{ws.max_row}"

    excel_table = Table(displayName=table_name, ref=table_ref)
    excel_table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(excel_table)

print(f"\n✅ Обработка завершена!")
print(f"Файл сохранён: {OUTPUT_FILE}")
print(f"Найдено ИП (включая ликвидированные): {df['Это_ИП'].sum()} из {len(df)}")