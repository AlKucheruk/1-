from openpyxl import load_workbook, Workbook
from copy import copy
import os
from datetime import datetime

file_name = "Выгрузка_Банка_20160101_20161230_prepared.xlsx"
wb = load_workbook(file_name)
ws = wb.active

folder_path = os.path.dirname(os.path.abspath(file_name))

# Создаём папки
months_folder = os.path.join(folder_path, "Месяцы")
quarters_folder = os.path.join(folder_path, "Кварталы")
half_years_folder = os.path.join(folder_path, "Полугодья")

os.makedirs(months_folder, exist_ok=True)
os.makedirs(quarters_folder, exist_ok=True)
os.makedirs(half_years_folder, exist_ok=True)

# Найти колонку с датой
header = [cell.value for cell in ws[1]]
try:
    date_col_index = header.index("Дата1С") + 1
except ValueError:
    print("Ошибка: колонка 'Дата1С' не найдена!")
    exit()

data_by_date = {}

for row in ws.iter_rows(min_row=2):
    cell = row[date_col_index - 1]
    date_value = cell.value

    if date_value is None:
        continue

    # Приводим дату к объекту datetime.date
    if isinstance(date_value, str):
        try:
            # Пробуем разные форматы дат
            if '.' in date_value:
                dt = datetime.strptime(date_value, '%d.%m.%Y')
            elif '-' in date_value:
                dt = datetime.strptime(date_value, '%Y-%m-%d')
            else:
                dt = datetime.strptime(date_value, '%d/%m/%Y')
            date_key = dt.date()
        except ValueError:
            print(f"Не удалось распознать дату: {date_value}")
            continue
    elif hasattr(date_value, 'date'):   # datetime или Timestamp
        date_key = date_value.date()
    else:
        date_key = date_value  # на всякий случай

    data_by_date.setdefault(date_key, []).append(row)

header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]

# Функция копирования ячейки со стилем
def copy_cell(source_cell, target_cell):
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

# Универсальная функция создания файла
def create_period_file(rows, target_folder, file_name):
    new_wb = Workbook()
    new_ws = new_wb.active

    # Заголовок
    for col_idx, cell in enumerate(header_row, start=1):
        new_cell = new_ws.cell(row=1, column=col_idx)
        copy_cell(cell, new_cell)

    # Данные
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, cell in enumerate(row, start=1):
            new_cell = new_ws.cell(row=row_idx, column=col_idx)
            copy_cell(cell, new_cell)

    new_file_path = os.path.join(target_folder, file_name)
    new_wb.save(new_file_path)

# Сортируем даты
sorted_dates = sorted(data_by_date.keys())

# ====================== ПОЛУГОДИЯ ======================
half_data = {}
for date in sorted_dates:
    # Защита от строк
    if isinstance(date, str):
        try:
            dt = datetime.strptime(date, '%Y-%m-%d').date() if '-' in date else datetime.strptime(date, '%d.%m.%Y').date()
            month = dt.month
        except:
            continue
    else:
        month = date.month

    half = 1 if month <= 6 else 2
    half_data.setdefault(half, []).extend(data_by_date[date])

for half, rows in half_data.items():
    new_file_name = f"Данные_Полугодие {half}.xlsx"
    create_period_file(rows, half_years_folder, new_file_name)

# ====================== КВАРТАЛЫ ======================
quarter_data = {}
for date in sorted_dates:
    if isinstance(date, str):
        try:
            dt = datetime.strptime(date, '%Y-%m-%d').date() if '-' in date else datetime.strptime(date, '%d.%m.%Y').date()
            month = dt.month
        except:
            continue
    else:
        month = date.month

    quarter = ((month - 1) // 3) + 1
    quarter_data.setdefault(quarter, []).extend(data_by_date[date])

for quarter, rows in quarter_data.items():
    new_file_name = f"Данные_Квартал {quarter}.xlsx"
    create_period_file(rows, quarters_folder, new_file_name)

# ====================== МЕСЯЦЫ ======================
month_data = {}
for date in sorted_dates:
    if isinstance(date, str):
        try:
            dt = datetime.strptime(date, '%Y-%m-%d').date() if '-' in date else datetime.strptime(date, '%d.%m.%Y').date()
            m = dt.month
        except:
            continue
    else:
        m = date.month

    month_data.setdefault(m, []).extend(data_by_date[date])

for month, rows in month_data.items():
    month_str = str(month).zfill(2)
    new_file_name = f"Данные_ {month_str}.xlsx"
    create_period_file(rows, months_folder, new_file_name)

print("Готово!")
print(f"Создано файлов:")
print(f"   • Полугодий: {len(half_data)}")
print(f"   • Кварталов: {len(quarter_data)}")
print(f"   • Месяцев:   {len(month_data)}")