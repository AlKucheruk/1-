# -*- coding: utf-8 -*-
# cspell:words openpyxl pydatetime sheetnames
"""
Сравнение выгрузок ЖурналДокументов_БП.xlsx и ЖурналДокументов_УТ.xlsx.

Исходный лист (первая страница) не меняется.

На второй странице «Результат» выводится «чистая» таблица: заголовок с первой
строки, без пустых строк сверху, плюс колонки «Статус в …» и «СуммаДокументаВ…»
(сумма найденной строки в противоположной выгрузке; для частичного совпадения
по дате — сумма строки УТ/БП с тем же ИНН, датой и направлением, ближайшая по сумме).
В файле БП после них дополнительно выводятся 8 колонок выгрузки УТ (с префиксом
«УТ.» в заголовке) — из той строки УТ, с которой выполнено сопоставление, и колонка
«Комментарий» с пояснением, в т.ч. при расхождениях.
Оформлена как умная таблица Excel: tbl_BP (файл БП) и tbl_UT (файл УТ).

Выгрузка БП может содержать 9 колонок (колонка «ВидОперации» между видом
документа и номером); на сравнение ключом она не влияет.

«Полное сходство» — в другой таблице есть строка с тем же ИНН, той же календарной
датой, той же суммой документа (±0.01) и тем же направлением («поступление» /
«списание»), см. classify_direction() по колонке вида документа.

«Полные» совпадения по ключу (ИНН+дата+сумма+направление) сопоставляются
по принципу 1:1 в порядке номеров строк: каждая строка УТ участвует не более
чем в одном полном ключа с БП, затем в частичных проверках участвуют только
её несопоставленные по полному ключу строки (и симметрично БП), чтобы единственный документ в УТ
не сравнивали повторно с чужими строками БП.

«Есть расхождения» — полного ключа нет, но есть строка с тем же ИНН, той же
датой и тем же направлением ИЛИ с тем же ИНН, той же суммой (±0.01) и тем же
направлением (направление по виду документа, см. classify_direction()).

«Строки нет» — нет ни полного совпадения, ни такого частичного совпадения.

Запуск из папки со скриптом:
  python compare_journals_bp_ut.py
"""

from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict
from collections.abc import Mapping, Sequence
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

BP_NAME = "ЖурналДокументов_БП.xlsx"
UT_NAME = "ЖурналДокументов_УТ.xlsx"
# Выгрузка БП после запроса с полем ВидОперации: 9 колонок до статуса включительно.
BP_SOURCE_DATA_COLS = 9
UT_SOURCE_DATA_COLS = 8
HEADER_ROW = 2
DATA_START_ROW = 3
RESULT_SHEET_NAME = "Результат"


def norm_inn(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return re.sub(r"\D", "", str(value).strip())


def norm_date(value: object) -> pd.Timestamp | None:
    dt = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        return None
    return pd.Timestamp(dt.normalize())


def norm_amount(value: object) -> float | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return None


def classify_direction(kind: object) -> str:
    """«поступление» | «списание» | «» по тексту вида документа."""
    if kind is None or (isinstance(kind, float) and pd.isna(kind)):
        return ""
    s = str(kind).strip()
    if not s:
        return ""
    low = s.lower()
    compact = "".join(low.split())

    def has_sub(*parts: str) -> bool:
        return any(p in low or p in compact for p in parts)

    # Сначала узкие исключения (корректировки)
    if has_sub("корректировкаприобретения", "корректировка приобретения"):
        return "поступление"
    if has_sub("корректировкареализации", "корректировка реализации"):
        return "списание"

    if has_sub("реализация", "отгрузк", "списание", "выдача"):
        return "списание"
    if has_sub("возврат", "поставщик"):
        return "списание"
    if has_sub("отчеткомитенту", "отчёткомитенту", "комитенту"):
        return "списание"

    if has_sub("приобретение", "поступление"):
        return "поступление"
    if has_sub("отклиента", "от клиента", "покупател", "возвраттоваровот"):
        return "поступление"
    if has_sub("отчеткомиссионера", "отчёткомиссионера", "комиссионера"):
        return "поступление"

    return ""


def load_table(path: Path) -> pd.DataFrame:
    return pd.read_excel(path, header=None)


def extract_dataframe(raw: pd.DataFrame, n_source_cols: int) -> pd.DataFrame:
    """
    n_source_cols=8 (УТ): дата, контрагент, ИНН, договор, вид документа, номер, сумма, статус.
    n_source_cols=9 (БП): то же + колонка «Вид операции» (ВидОперации) между видом документа и номером.
    """
    body = raw.iloc[DATA_START_ROW:, :n_source_cols].copy()
    body.columns = list(range(n_source_cols))
    if n_source_cols == 8:
        rename = {
            0: "date",
            1: "client",
            2: "inn",
            3: "contract",
            4: "kind",
            5: "num",
            6: "amount",
            7: "status",
        }
    elif n_source_cols == 9:
        rename = {
            0: "date",
            1: "client",
            2: "inn",
            3: "contract",
            4: "kind",
            5: "op_kind",
            6: "num",
            7: "amount",
            8: "status",
        }
    else:
        raise ValueError("n_source_cols must be 8 or 9")
    return body.rename(columns=rename).reset_index(drop=True)


def _row_full_key_tuple(r: pd.Series) -> tuple | None:
    """Полный ключ сравнения или None, если строку нельзя классифицировать."""
    inn = norm_inn(r.get("inn"))
    d = norm_date(r.get("date"))
    a = norm_amount(r.get("amount"))
    dr = classify_direction(r.get("kind"))
    if inn == "" or d is None or a is None or not dr:
        return None
    return (inn, d, a, dr)


def allocate_full_key_pairs(
    bp_body: pd.DataFrame, ut_body: pd.DataFrame
) -> tuple[dict[int, int], dict[int, int]]:
    """
    Однозначное сопоставление по полному ключу: по каждому ключу сортируем
    номера строк БП и УТ и сшиваем по min(число в БП, число в УТ). Строка УТ
    не участвует в двух полных парах — дальше она исключается из частичных правил.
    """
    by_key_bp: defaultdict[tuple, list[int]] = defaultdict(list)
    by_key_ut: defaultdict[tuple, list[int]] = defaultdict(list)
    for i in range(len(bp_body)):
        fk = _row_full_key_tuple(bp_body.iloc[i])
        if fk is not None:
            by_key_bp[fk].append(i)
    for j in range(len(ut_body)):
        fk = _row_full_key_tuple(ut_body.iloc[j])
        if fk is not None:
            by_key_ut[fk].append(j)
    bp_to_ut: dict[int, int] = {}
    ut_to_bp: dict[int, int] = {}
    for fk in set(by_key_bp) | set(by_key_ut):
        bl = sorted(by_key_bp.get(fk, []))
        ul = sorted(by_key_ut.get(fk, []))
        for k in range(min(len(bl), len(ul))):
            bi, uj = bl[k], ul[k]
            bp_to_ut[bi] = uj
            ut_to_bp[uj] = bi
    return bp_to_ut, ut_to_bp


def _iter_body_row_indices(body: pd.DataFrame, row_indices: set[int] | None) -> list[int]:
    if row_indices is None:
        return list(range(len(body)))
    return sorted(row_indices)


def build_indexes(body: pd.DataFrame, row_indices: set[int] | None = None):
    full_keys: set[tuple] = set()
    by_inn_date_dir: set[tuple] = set()
    by_inn_amt_dir: set[tuple] = set()

    for i in _iter_body_row_indices(body, row_indices):
        r = body.iloc[i]
        inn = norm_inn(r.get("inn"))
        d = norm_date(r.get("date"))
        a = norm_amount(r.get("amount"))
        dr = classify_direction(r.get("kind"))
        if inn == "" or d is None or a is None or not dr:
            continue
        full_keys.add((inn, d, a, dr))
        by_inn_date_dir.add((inn, d, dr))
        by_inn_amt_dir.add((inn, a, dr))

    return full_keys, by_inn_date_dir, by_inn_amt_dir


def build_amount_buckets(
    body: pd.DataFrame,
    row_indices: set[int] | None = None,
) -> tuple[
    defaultdict[tuple, list[float]],
    defaultdict[tuple, list[float]],
    defaultdict[tuple, list[float]],
]:
    """
    Списки сумм по ключам противоположной таблицы (для подбора «суммы в другой базе»).
    row_indices: если задан, учитывать только эти номера строк (например свободные УТ/БП).
    """
    by_full: defaultdict[tuple, list[float]] = defaultdict(list)
    by_inn_date_dir: defaultdict[tuple, list[float]] = defaultdict(list)
    by_inn_amt_dir: defaultdict[tuple, list[float]] = defaultdict(list)
    for i in _iter_body_row_indices(body, row_indices):
        r = body.iloc[i]
        inn = norm_inn(r.get("inn"))
        d = norm_date(r.get("date"))
        a = norm_amount(r.get("amount"))
        dr = classify_direction(r.get("kind"))
        if inn == "" or d is None or a is None or not dr:
            continue
        by_full[(inn, d, a, dr)].append(a)
        by_inn_date_dir[(inn, d, dr)].append(a)
        by_inn_amt_dir[(inn, a, dr)].append(a)
    return by_full, by_inn_date_dir, by_inn_amt_dir


def build_row_index_maps(
    body: pd.DataFrame,
    row_indices: set[int] | None = None,
) -> tuple[
    defaultdict[tuple, list[int]],
    defaultdict[tuple, list[int]],
    defaultdict[tuple, list[int]],
]:
    """
    Номера строк (0..len-1) по тем же ключам, что build_amount_buckets,
    без полного перебора opposite_body на каждую строку БП.
    """
    by_row_full: defaultdict[tuple, list[int]] = defaultdict(list)
    by_row_inn_date_dir: defaultdict[tuple, list[int]] = defaultdict(list)
    by_row_inn_amt_dir: defaultdict[tuple, list[int]] = defaultdict(list)
    for i in _iter_body_row_indices(body, row_indices):
        r = body.iloc[i]
        inn = norm_inn(r.get("inn"))
        d = norm_date(r.get("date"))
        a = norm_amount(r.get("amount"))
        dr = classify_direction(r.get("kind"))
        if inn == "" or d is None or a is None or not dr:
            continue
        by_row_full[(inn, d, a, dr)].append(i)
        by_row_inn_date_dir[(inn, d, dr)].append(i)
        by_row_inn_amt_dir[(inn, a, dr)].append(i)
    return by_row_full, by_row_inn_date_dir, by_row_inn_amt_dir


def _pair_row_to_opposite(
    row: pd.Series,
    opposite_body: pd.DataFrame,
    by_full: Mapping[tuple, Sequence[float]],
    by_inn_date_dir: Mapping[tuple, Sequence[float]],
    by_inn_amt_dir: Mapping[tuple, Sequence[float]],
    row_by_full: Mapping[tuple, Sequence[int]],
    row_by_inn_date_dir: Mapping[tuple, Sequence[int]],
    row_by_inn_amt_dir: Mapping[tuple, Sequence[int]],
) -> tuple[int | None, str]:
    """
    Индекс строки в opposite_body, выбранной по той же логике, что и сумма
    в opposite_amount_from_side, и способ: full | date_dir | amt_dir | none.
    """
    inn = norm_inn(row.get("inn"))
    d = norm_date(row.get("date"))
    a = norm_amount(row.get("amount"))
    dr = classify_direction(row.get("kind"))
    if inn == "" or d is None or a is None or not dr:
        return None, "none"
    fk = (inn, d, a, dr)
    if fk in by_full and by_full[fk] and row_by_full.get(fk):
        return int(row_by_full[fk][0]), "full"
    if (inn, d, dr) in by_inn_date_dir:
        amount_options = by_inn_date_dir[(inn, d, dr)]
        if not amount_options:
            return None, "none"
        best_amt = float(min(amount_options, key=lambda x: abs(x - a)))
        for i in row_by_inn_date_dir.get((inn, d, dr), ()):
            ua = norm_amount(opposite_body.iloc[i].get("amount"))
            if ua is not None and abs(ua - best_amt) < 0.01:
                return int(i), "date_dir"
        return None, "none"
    if (inn, a, dr) in by_inn_amt_dir and row_by_inn_amt_dir.get((inn, a, dr)):
        return int(row_by_inn_amt_dir[(inn, a, dr)][0]), "amt_dir"
    return None, "none"


def opposite_amount_from_side(
    row: pd.Series,
    by_full: Mapping[tuple, Sequence[float]],
    by_inn_date_dir: Mapping[tuple, Sequence[float]],
    by_inn_amt_dir: Mapping[tuple, Sequence[float]],
) -> float | None:
    """
    Сумма документа на противоположной стороне для строки row.
    Полное совпадение ключа — сумма из полного ключа (совпадает с суммой строки).
    Иначе при совпадении ИНН+дата+направление — среди строк с тем же ИНН, датой
    и направлением берётся сумма, минимально отличающаяся от суммы текущей строки.
    Иначе при совпадении ИНН+сумма+направление — та же сумма.
    """
    inn = norm_inn(row.get("inn"))
    d = norm_date(row.get("date"))
    a = norm_amount(row.get("amount"))
    dr = classify_direction(row.get("kind"))
    if inn == "" or d is None or a is None or not dr:
        return None
    fk = (inn, d, a, dr)
    if fk in by_full and by_full[fk]:
        return float(by_full[fk][0])
    if (inn, d, dr) in by_inn_date_dir:
        amount_options = by_inn_date_dir[(inn, d, dr)]
        if not amount_options:
            return None
        return float(min(amount_options, key=lambda x: abs(x - a)))
    if (inn, a, dr) in by_inn_amt_dir:
        amount_options = by_inn_amt_dir[(inn, a, dr)]
        return float(amount_options[0]) if amount_options else float(a)
    return None


def _format_date_ru(d: pd.Timestamp | None) -> str:
    if d is None:
        return "—"
    return d.strftime("%d.%m.%Y")


def _format_amt(x: object) -> str:
    a = norm_amount(x)
    if a is None:
        return "—"
    return f"{a:,.2f}".replace(",", " ")


def comment_for_bp_result(
    status: str,
    r_bp: pd.Series,
    ut_idx: int | None,
    pair_mode: str,
    ut_body: pd.DataFrame,
) -> str:
    """Пояснение к строке БП; не меняет текст колонки «Статус в УТ»."""
    inn = norm_inn(r_bp.get("inn"))
    d = norm_date(r_bp.get("date"))
    a = norm_amount(r_bp.get("amount"))
    dr = classify_direction(r_bp.get("kind"))
    if status.startswith("нет данных"):
        parts: list[str] = []
        if inn == "":
            parts.append("нет ИНН")
        if d is None:
            parts.append("нет или неверная дата")
        if a is None:
            parts.append("нет или неверная сумма")
        if not dr:
            parts.append("не определено направление по виду документа")
        if parts:
            return "Нельзя сравнить: " + "; ".join(parts) + "."
        return "Нельзя сравнить с УТ: не хватает обязательных полей."
    if status == "строки нет":
        return (
            "В выгрузке УТ нет частично подходящей строки с тем же ИНН и "
            "(той же датой и направлением или той же суммой и направлением); "
            "полного совпадения по ИНН+дате+сумме+направлению тоже нет."
        )
    if ut_idx is None or ut_idx >= len(ut_body):
        if status == "есть расхождения":
            return "Статус «есть расхождения», но не удалось сопоставить конкретную строку УТ (проверьте дубли/данные)."
        if status == "полное сходство":
            return "Полное совпадение по правилам, но не найдена соответствующая строка в таблице УТ (проверьте дубли/данные)."
        return ""
    u = ut_body.iloc[ut_idx]
    ua = norm_amount(u.get("amount"))
    ud = norm_date(u.get("date"))
    kind_text_ut = u.get("kind")
    kind_text_bp = r_bp.get("kind")
    diff_amt: str | None = None
    if a is not None and ua is not None and abs(a - ua) > 0.01:
        diff_amt = f"в БП {_format_amt(a)}, в сопоставленной строке УТ {_format_amt(ua)} (разница {_format_amt(abs(a - ua))} руб.)."

    if status == "полное сходство":
        return (
            "Совпадение: ИНН, дата, сумма, направление (по виду документа) с отображаемой строкой УТ. "
        ).strip()
    if status != "есть расхождения":
        return ""

    if pair_mode == "date_dir":
        lines = [
            "Частичное сопоставление: тот же ИНН, та же дата и направление, что в выбранной строке УТ; сумма отличается (полного ключа нет).",
        ]
        if diff_amt:
            lines.append(diff_amt)
        if kind_text_bp is not None and kind_text_ut is not None and str(kind_text_bp).strip() != str(kind_text_ut).strip():
            lines.append(
                f"По тексту «вид документа»: в БП «{kind_text_bp!s}», в УТ «{kind_text_ut!s}»."
            )
        return " ".join(lines)
    if pair_mode == "amt_dir":
        lines = [
            "Частичное сопоставление: тот же ИНН, та же сумма и направление, что в выбранной строке УТ; даты могут не совпадать.",
        ]
        if d is not None and ud is not None and d != ud:
            lines.append(f"Дата: в БП {_format_date_ru(d)}, в УТ {_format_date_ru(ud)}.")
        if kind_text_bp is not None and kind_text_ut is not None and str(kind_text_bp).strip() != str(kind_text_ut).strip():
            lines.append(
                f"Вид документа: в БП «{kind_text_bp!s}», в УТ «{kind_text_ut!s}»."
            )
        return " ".join(lines)
    return (
        "Полного совпадения нет; по правилам подобрана строка УТ (см. ячейки слева). "
    ).strip()


def status_bp_vs_ut(
    r: pd.Series,
    ut_full: set[tuple],
    ut_id_dir: set[tuple],
    ut_ia_dir: set[tuple],
) -> str:
    inn = norm_inn(r.get("inn"))
    d = norm_date(r.get("date"))
    a = norm_amount(r.get("amount"))
    dr = classify_direction(r.get("kind"))

    if inn == "" or d is None or a is None or not dr:
        return "нет данных для сравнения (ИНН/дата/сумма/направление)"

    if (inn, d, a, dr) in ut_full:
        return "полное сходство"

    if (inn, d, dr) in ut_id_dir or (inn, a, dr) in ut_ia_dir:
        return "есть расхождения"
    return "строки нет"


def status_ut_vs_bp(
    r: pd.Series,
    bp_full: set[tuple],
    bp_id_dir: set[tuple],
    bp_ia_dir: set[tuple],
) -> str:
    inn = norm_inn(r.get("inn"))
    d = norm_date(r.get("date"))
    a = norm_amount(r.get("amount"))
    dr = classify_direction(r.get("kind"))

    if inn == "" or d is None or a is None or not dr:
        return "нет данных для сравнения (ИНН/дата/сумма/направление)"

    if (inn, d, a, dr) in bp_full:
        return "полное сходство"

    if (inn, d, dr) in bp_id_dir or (inn, a, dr) in bp_ia_dir:
        return "есть расхождения"
    return "строки нет"


def _excel_cell(value: object) -> object:
    """Значение для ячейки openpyxl (без NaN/NaT в «сыром» виде)."""
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def ut8_headers_from_raw(raw_ut: pd.DataFrame) -> list[str]:
    """Восемь заголовков колонок листа УТ с префиксом «УТ.»."""
    m = min(8, int(raw_ut.shape[1]))
    out: list[str] = []
    for j in range(8):
        if j < m:
            h = raw_ut.iat[HEADER_ROW, j]
            if isinstance(h, float) and pd.isna(h):
                h = ""
            h = str(h).strip()
            if h:
                out.append(f"УТ. {h}")
            else:
                out.append(f"УТ. кол. {j + 1}")
        else:
            out.append(f"УТ. кол. {j + 1}")
    return out


def write_result_smart_table(
    workbook_path: Path,
    raw: pd.DataFrame,
    statuses: list[str],
    opposite_amounts: list[float | None],
    status_header: str,
    opposite_sum_header: str,
    table_display_name: str,
    n_source_cols: int,
    *,
    ut_eight_headers: list[str] | None = None,
    ut_eight_rows: list[list[object]] | None = None,
    comments: list[str] | None = None,
) -> None:
    """
    Добавляет/обновляет лист RESULT_SHEET_NAME (вторая позиция в книге):
    строка 1 — заголовки из исходной строки HEADER_ROW + status_header
    + opposite_sum_header;
    далее данные из колонок 0..n_source_cols-1 исходного листа + статус + сумма
    в противоположной базе (если есть);
    опционально: 8 колонок снимка УТ (книга БП) и колонка «Комментарий»;
    оформление — таблица Excel с именем table_display_name (tbl_BP / tbl_UT).
    """
    wb = load_workbook(workbook_path, read_only=False, data_only=False)
    if RESULT_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[RESULT_SHEET_NAME])
    ws = wb.create_sheet(RESULT_SHEET_NAME, 1)

    n = len(statuses)
    use_ut8 = bool(ut_eight_headers and ut_eight_rows is not None)
    use_com = bool(comments)
    if use_ut8 and (len(ut_eight_rows or []) != n or len(ut_eight_headers) != 8):
        raise ValueError("ut_eight: need 8 headers and n rows of 8 values each")
    if use_com and len(comments or []) != n:
        raise ValueError("comments: length must match number of data rows")
    header: list = []
    for j in range(n_source_cols):
        h = raw.iat[HEADER_ROW, j]
        if isinstance(h, float) and pd.isna(h):
            h = ""
        header.append(h)
    header.append(status_header)
    header.append(opposite_sum_header)
    if use_ut8:
        header.extend(ut_eight_headers or [])
    if use_com:
        header.append("Комментарий")

    ws.append([_excel_cell(x) for x in header])
    for i in range(n):
        row = [_excel_cell(raw.iat[DATA_START_ROW + i, j]) for j in range(n_source_cols)]
        row.append(statuses[i])
        oa = opposite_amounts[i]
        row.append(round(float(oa), 2) if oa is not None else None)
        if use_ut8 and ut_eight_rows is not None:
            row.extend(_excel_cell(x) for x in ut_eight_rows[i])
        if use_com and comments is not None:
            row.append(_excel_cell(comments[i]))
        ws.append(row)

    ncols = len(header)
    last_row = 1 + n
    last_col = get_column_letter(ncols)
    ref = f"A1:{last_col}{last_row}"
    tab = Table(displayName=table_display_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)
    wb.save(workbook_path)


def _print_error(msg: str) -> None:
    """Сообщение в stderr и в stdout — чтобы вывод «Run» в IDE не казался пустым."""
    print(msg, file=sys.stderr, flush=True)
    print(msg, flush=True)


def main() -> int:
    base = Path(__file__).resolve().parent
    bp_path = base / BP_NAME
    ut_path = base / UT_NAME
    print(
        f"Сравнение журналов: папка данных — {base}",
        flush=True,
    )
    if not bp_path.is_file() or not ut_path.is_file():
        _print_error(
            f"Нужны файлы «{BP_NAME}» и «{UT_NAME}» в папке:\n{base}\n"
            "(без этих выгрузок сравнение не выполняется)."
        )
        return 1

    raw_bp = load_table(bp_path)
    raw_ut = load_table(ut_path)
    if raw_bp.shape[1] < 8 or raw_ut.shape[1] < 8:
        _print_error("Error: each xlsx needs at least 8 data columns.")
        return 1
    bp_cols = min(BP_SOURCE_DATA_COLS, raw_bp.shape[1])
    ut_cols = min(UT_SOURCE_DATA_COLS, raw_ut.shape[1])
    if bp_cols < BP_SOURCE_DATA_COLS:
        print(
            "Warning: BP xlsx has",
            bp_cols,
            "columns (expected",
            str(BP_SOURCE_DATA_COLS) + " after query update); using",
            bp_cols,
            "for compare.",
            file=sys.stderr,
        )
    bp_body = extract_dataframe(raw_bp, bp_cols)
    ut_body = extract_dataframe(raw_ut, ut_cols)
    print(
        f"Загружено: БП {len(bp_body)} строк, УТ {len(ut_body)} строк.",
        flush=True,
    )

    n_bp = len(bp_body)
    n_ut = len(ut_body)
    bp_to_ut, ut_to_bp = allocate_full_key_pairs(bp_body, ut_body)
    ut_free: set[int] = set(range(n_ut)) - set(ut_to_bp)
    bp_free: set[int] = set(range(n_bp)) - set(bp_to_ut)
    print(
        f"Полных пар по ключу: {len(bp_to_ut)}; "
        f"свободно для частичного сопоставления — БП {len(bp_free)}, УТ {len(ut_free)}.",
        flush=True,
    )

    # Частичные правила: только по строкам, не занятым в полных парах
    ut_full, ut_id, ut_ia = build_indexes(ut_body, ut_free)
    ut_am_full, ut_am_id_dir, ut_am_ia_dir = build_amount_buckets(ut_body, ut_free)
    ut_row_full, ut_row_id_dir, ut_row_amt_dir = build_row_index_maps(ut_body, ut_free)

    bp_full, bp_id, bp_ia = build_indexes(bp_body, bp_free)
    bp_am_full, bp_am_id_dir, bp_am_ia_dir = build_amount_buckets(bp_body, bp_free)

    bp_statuses: list[str] = []
    bp_opp_amounts: list[float | None] = []
    bp_ut8_rows: list[list[object]] = []
    bp_comments: list[str] = []
    h_ut8 = ut8_headers_from_raw(raw_ut)

    for i in range(n_bp):
        r = bp_body.iloc[i]
        if i in bp_to_ut:
            st = "полное сходство"
            uj = bp_to_ut[i]
            ut_i: int | None = uj
            match_mode = "full"
            bp_opp_amounts.append(norm_amount(ut_body.iloc[uj].get("amount")))
        else:
            st = status_bp_vs_ut(r, ut_full, ut_id, ut_ia)
            bp_opp_amounts.append(
                opposite_amount_from_side(r, ut_am_full, ut_am_id_dir, ut_am_ia_dir)
            )
            ut_i, match_mode = _pair_row_to_opposite(
                r,
                ut_body,
                ut_am_full,
                ut_am_id_dir,
                ut_am_ia_dir,
                ut_row_full,
                ut_row_id_dir,
                ut_row_amt_dir,
            )
        row8: list[object] = [None] * 8
        n_ut_raw = min(8, int(raw_ut.shape[1]))
        if ut_i is not None and 0 <= ut_i < n_ut:
            for j in range(n_ut_raw):
                row8[j] = raw_ut.iat[DATA_START_ROW + ut_i, j]
        bp_statuses.append(st)
        bp_ut8_rows.append(row8)
        bp_comments.append(comment_for_bp_result(st, r, ut_i, match_mode, ut_body))

    print("Строки БП обработаны, сравниваем УТ с БП…", flush=True)

    ut_statuses: list[str] = []
    ut_opp_amounts: list[float | None] = []
    for j in range(n_ut):
        if j in ut_to_bp:
            ut_statuses.append("полное сходство")
            bi = ut_to_bp[j]
            ut_opp_amounts.append(norm_amount(bp_body.iloc[bi].get("amount")))
        else:
            urow = ut_body.iloc[j]
            ut_statuses.append(status_ut_vs_bp(urow, bp_full, bp_id, bp_ia))
            ut_opp_amounts.append(
                opposite_amount_from_side(
                    urow, bp_am_full, bp_am_id_dir, bp_am_ia_dir
                )
            )

    print("Запись листа «Результат» в Excel…", flush=True)
    write_result_smart_table(
        bp_path,
        raw_bp,
        bp_statuses,
        bp_opp_amounts,
        "Статус в УТ",
        "СуммаДокументаВУТ",
        "tbl_BP",
        bp_cols,
        ut_eight_headers=h_ut8,
        ut_eight_rows=bp_ut8_rows,
        comments=bp_comments,
    )
    write_result_smart_table(
        ut_path,
        raw_ut,
        ut_statuses,
        ut_opp_amounts,
        "Статус в БП",
        "СуммаДокументаВБП",
        "tbl_UT",
        ut_cols,
    )

    print("OK. Сравнение завершено (файлы записаны).", flush=True)
    print(" ", bp_path.name, "- sheet", RESULT_SHEET_NAME, "table tbl_BP")
    print(" ", ut_path.name, "- sheet", RESULT_SHEET_NAME, "table tbl_UT")
    print("\nSummary BP:")
    for k, v in Counter(bp_statuses).most_common():
        print(" ", v, repr(k))
    print("Summary UT:")
    for k, v in Counter(ut_statuses).most_common():
        print(" ", v, repr(k))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
