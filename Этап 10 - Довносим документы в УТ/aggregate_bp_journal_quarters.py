# cspell:words openpyxl prihod rashod
"""
Читает `ЖурналДокументов_БП обработанный.xlsx`, лист Source, tbl_BP8 (колонки A:U).
Добавляет/обновляет листы «Приход» и «Расход» с квартальной агрегацией.

Приход: все строки, у которых в «ВидДокумента» встречается слово «Поступление»
(в т.ч. «Поступление (акт…)», «Поступление доп. расходов», …).

Расход: только «Реализация».

Ключ группировки: (ИНН, Статья, год, квартал). «СуммаДокумента» — сумма.
Колонка «Дата» в результате: 2-е число 3-го месяца квартала (март, июнь, сентябрь, декабрь).
"""
from __future__ import annotations

import re
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SOURCE_SHEET = "Source"
INPUT_NAME = "ЖурналДокументов_БП обработанный.xlsx"
SHEET_PRIHOD = "Приход"
SHEET_RASHOD = "Расход"
OUT_COLS = ["Дата", "Контрагент", "ИНН", "Статья", "СуммаДокумента"]
TABLE_PRIHOD = "tbl_Prihod_Q"
TABLE_RASHOD = "tbl_Rashod_Q"

# Квартал → 2-й день 3-го месяца (март, июнь, сентябрь, декабрь)
_QUARTER_TO_MONTH = {1: 3, 2: 6, 3: 9, 4: 12}


def _quarter_label_day(year: int, quarter: int) -> date:
    m = _QUARTER_TO_MONTH[quarter]
    return date(int(year), m, 2)


def _norm_inn(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _read_source(path: Path) -> pd.DataFrame:
    df = pd.read_excel(
        path,
        sheet_name=SOURCE_SHEET,
        usecols="A:U",
        engine="openpyxl",
    )
    need = {"Дата", "Контрагент", "ИНН", "Статья", "ВидДокумента", "СуммаДокумента"}
    missing = need - set(df.columns)
    if missing:
        raise SystemExit(f"В таблице нет колонок: {sorted(missing)}")
    return df


def _aggregate(
    df: pd.DataFrame,
    mask: pd.Series,
) -> pd.DataFrame:
    sub = df.loc[mask].copy()
    sub["_inn"] = sub["ИНН"].map(_norm_inn)
    sub["_d"] = pd.to_datetime(sub["Дата"], errors="coerce", dayfirst=True)
    sub = sub[sub["_d"].notna()]
    if sub.empty:
        return pd.DataFrame(columns=OUT_COLS)

    sub["СуммаДокумента"] = pd.to_numeric(sub["СуммаДокумента"], errors="coerce").fillna(0.0)
    sub["Статья"] = sub["Статья"].fillna("").map(lambda t: str(t).strip() if t is not None else "")

    sub["_y"] = sub["_d"].dt.year
    sub["_q"] = sub["_d"].dt.quarter

    g = sub.groupby(["_inn", "Статья", "_y", "_q"], dropna=False, as_index=False)
    agg = g.agg(
        СуммаДокумента=("СуммаДокумента", "sum"),
        Контрагент=("Контрагент", "first"),
    )

    def _row_date(r) -> date:
        return _quarter_label_day(r["_y"], r["_q"])

    agg["Дата"] = agg.apply(_row_date, axis=1)
    agg["ИНН"] = agg["_inn"]

    out = agg[OUT_COLS].copy()
    out["СуммаДокумента"] = out["СуммаДокумента"].round(2)
    out = out.sort_values(["Дата", "ИНН", "Статья"], kind="mergesort").reset_index(drop=True)
    return out


def _mask_prihod(ser: pd.Series) -> pd.Series:
    s = ser.fillna("").astype(str)
    return s.str.contains("Поступление", case=False, regex=False, na=False)


def _mask_rashod(ser: pd.Series) -> pd.Series:
    return ser.fillna("").astype(str).str.strip() == "Реализация"


def _remove_sheet_if_exists(wb, name: str) -> None:
    if name in wb.sheetnames:
        wb.remove(wb[name])


def _add_table(wb, sheet_name: str, table_name: str) -> None:
    ws = wb[sheet_name]
    nrows = ws.max_row
    ncols = 5
    if nrows < 1:
        return
    ref = f"A1:{get_column_letter(ncols)}{nrows}"
    # displayName: только буквы/цифры/подчёркивания, с буквы
    safe = re.sub(r"[^A-Za-z0-9_]", "_", table_name)
    if not safe or not safe[0].isalpha():
        safe = "T_" + safe
    tab = Table(displayName=safe[:255], name=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def main() -> None:
    base = Path(__file__).resolve().parent
    path = base / INPUT_NAME
    if not path.is_file():
        print(f"Файл не найден: {path}", file=sys.stderr)
        sys.exit(1)

    df = _read_source(path)
    d_prih = _aggregate(df, _mask_prihod(df["ВидДокумента"]))
    d_rash = _aggregate(df, _mask_rashod(df["ВидДокумента"]))

    wb = load_workbook(path)
    _remove_sheet_if_exists(wb, SHEET_PRIHOD)
    _remove_sheet_if_exists(wb, SHEET_RASHOD)
    wb.save(path)
    wb.close()

    with pd.ExcelWriter(  # noqa: PD901
        path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace",
    ) as writer:
        d_prih.to_excel(writer, sheet_name=SHEET_PRIHOD, index=False)
        d_rash.to_excel(writer, sheet_name=SHEET_RASHOD, index=False)

    wb2 = load_workbook(path)
    _add_table(wb2, SHEET_PRIHOD, TABLE_PRIHOD)
    _add_table(wb2, SHEET_RASHOD, TABLE_RASHOD)
    wb2.save(path)
    wb2.close()

    print(
        f"OK. {SHEET_PRIHOD}: {len(d_prih)} строк; {SHEET_RASHOD}: {len(d_rash)} строк. Файл: {path.name}"
    )


if __name__ == "__main__":
    main()
