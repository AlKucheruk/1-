import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from datetime import datetime, date
import tkinter as tk
from tkinter import filedialog, messagebox

import pandas as pd


SCRIPT_DIR = Path(__file__).resolve().parent
MAPPING_FILE = SCRIPT_DIR / "mapping_operations_ut.xlsx"
DDS_MAPPING_FILE = SCRIPT_DIR / "dds_mapping_ut_bp.xlsx"

EXPECTED_COLUMNS = [
    "УИД",
    "ТипДокумента",
    "Номер1С",
    "Дата1С",
    "НомерБанк",
    "ДатаБанк",
    "Организация",
    "ИНН_Орг",
    "КПП_Орг",
    "НашСчет",
    "НашБанк",
    "НашБИК",
    "ВидОперации",
    "Контрагент",
    "ИНН_Контр",
    "КПП_Контр",
    "КодКонтр",
    "Договор",
    "ДокОснование",
    "Сумма",
    "Валюта",
    "Назначение",
    "СчетРасчетов",
    "СчетКонтр",
    "БанкКонтр",
    "БИК_Контр",
    "СтатьяДДС",
    "Комментарий",
    "Ответственный",
]

MAPPING_REQUIRED_COLUMNS = [
    "ТипДокументаБП",
    "ВидОперацииБП",
    "ДокументУТ",
    "ОперацияУТ",
    "ГруппаУчета",
    "НаправлениеДенег",
    "ТребуетПроверки",
    "КомментарийМаппинга",
]

DDS_MAPPING_REQUIRED_COLUMNS = [
    "ТипДокументаБП",
    "ВидОперацииБП",
    "СтатьяДДС_УТ",
    "ПапкаДДС",
    "НаправлениеДенег",
    "ТребуетПроверки",
    "Комментарий",
]


def normalize_text(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value)
    text = text.replace("\xa0", " ")
    text = text.replace("\u202f", " ")
    text = text.strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_key(value) -> str:
    return normalize_text(value).lower()


def ask_file_path() -> Path:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    file_path = filedialog.askopenfilename(
        title="Выберите Excel-файл банковской выгрузки из БП",
        filetypes=[
            ("Excel files", "*.xlsx *.xls"),
            ("All files", "*.*")
        ]
    )

    root.destroy()

    if not file_path:
        raise ValueError("Файл не выбран.")

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")

    return path


def load_excel_first_sheet(path: Path) -> pd.DataFrame:
    xls = pd.ExcelFile(path)
    if not xls.sheet_names:
        raise ValueError("В Excel-файле не найдено ни одного листа.")
    sheet_name = xls.sheet_names[0]
    return pd.read_excel(path, sheet_name=sheet_name)


def validate_bank_file(df: pd.DataFrame):
    missing = [col for col in EXPECTED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(
            "Файл не соответствует ожидаемому формату.\n"
            f"Не хватает колонок: {missing}"
        )


def validate_mapping_file(df_mapping: pd.DataFrame):
    missing = [col for col in MAPPING_REQUIRED_COLUMNS if col not in df_mapping.columns]
    if missing:
        raise ValueError(
            "Файл маппинга операций не соответствует ожидаемому формату.\n"
            f"Не хватает колонок: {missing}"
        )


def validate_dds_mapping_file(df_mapping: pd.DataFrame):
    missing = [col for col in DDS_MAPPING_REQUIRED_COLUMNS if col not in df_mapping.columns]
    if missing:
        raise ValueError(
            "Файл маппинга статей ДДС не соответствует ожидаемому формату.\n"
            f"Не хватает колонок: {missing}"
        )


def load_mapping(mapping_path: Path) -> pd.DataFrame:
    if not mapping_path.exists():
        raise FileNotFoundError(
            f"Не найден файл маппинга операций: {mapping_path}"
        )

    xls = pd.ExcelFile(mapping_path)
    sheet_name = "mapping" if "mapping" in xls.sheet_names else xls.sheet_names[0]
    df_mapping = pd.read_excel(mapping_path, sheet_name=sheet_name)

    validate_mapping_file(df_mapping)

    df_mapping = df_mapping.copy()
    df_mapping["ТипДокументаБП_norm"] = df_mapping["ТипДокументаБП"].apply(normalize_key)
    df_mapping["ВидОперацииБП_norm"] = df_mapping["ВидОперацииБП"].apply(normalize_key)

    return df_mapping


def load_dds_mapping(mapping_path: Path) -> pd.DataFrame:
    if not mapping_path.exists():
        raise FileNotFoundError(
            f"Не найден файл маппинга статей ДДС: {mapping_path}"
        )

    xls = pd.ExcelFile(mapping_path)
    sheet_name = "mapping" if "mapping" in xls.sheet_names else xls.sheet_names[0]
    df_mapping = pd.read_excel(mapping_path, sheet_name=sheet_name)

    validate_dds_mapping_file(df_mapping)

    df_mapping = df_mapping.copy()
    df_mapping["ТипДокументаБП_norm"] = df_mapping["ТипДокументаБП"].apply(normalize_key)
    df_mapping["ВидОперацииБП_norm"] = df_mapping["ВидОперацииБП"].apply(normalize_key)

    return df_mapping


def parse_document_date(value):
    if pd.isna(value):
        return None

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime().date()

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = normalize_text(value)
    if not text:
        return None

    formats = [
        "%Y-%m-%d",
        "%d.%m.%Y",
        "%d.%m.%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M:%S",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    dt = pd.to_datetime(text, errors="coerce", dayfirst=True)
    if pd.notna(dt):
        return dt.date()

    return None


def normalize_date_for_excel(value):
    parsed = parse_document_date(value)
    if parsed is None:
        return pd.NaT
    return pd.Timestamp(parsed)


def normalize_inn(value) -> str:
    if pd.isna(value):
        return ""

    if isinstance(value, (int,)):
        return str(value)

    if isinstance(value, float):
        if pd.isna(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return format(Decimal(str(value)).quantize(Decimal("1")), "f")

    text = str(value).strip()
    if not text:
        return ""

    text = text.replace("\xa0", "")
    text = text.replace("\u202f", "")
    text = text.replace(" ", "")

    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".")[0]

    sci_candidate = text.replace(",", ".")
    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?[eE][+-]?\d+", sci_candidate):
        try:
            dec = Decimal(sci_candidate)
            text = format(dec.quantize(Decimal("1")), "f")
        except (InvalidOperation, ValueError):
            pass

    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".")[0]

    if "." in text and re.fullmatch(r"\d+\.\d+", text):
        left, right = text.split(".", 1)
        if set(right) == {"0"}:
            text = left

    text = re.sub(r"[^\d]", "", text)
    return text


def normalize_kpp(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    text = text.replace("\xa0", "")
    text = text.replace("\u202f", "")
    text = text.replace(" ", "")
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".")[0]
    text = re.sub(r"[^\d]", "", text)
    return text


def normalize_account(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    text = text.replace("\xa0", "")
    text = text.replace("\u202f", "")
    text = text.replace(" ", "")

    sci_candidate = text.replace(",", ".")
    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?[eE][+-]?\d+", sci_candidate):
        try:
            dec = Decimal(sci_candidate)
            text = format(dec.quantize(Decimal("1")), "f")
        except (InvalidOperation, ValueError):
            pass

    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".")[0]

    text = re.sub(r"[^\d]", "", text)
    return text


def normalize_bik(value) -> str:
    return normalize_account(value)


def parse_amount(value):
    if pd.isna(value):
        return None

    if isinstance(value, (int, float)) and not pd.isna(value):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    text = text.replace("\xa0", "")
    text = text.replace("\u202f", "")
    text = text.replace(" ", "")

    if "," in text and "." in text:
        text = text.replace(" ", "")
        text = text.replace(",", ".")
    else:
        text = text.replace(",", ".")

    text = re.sub(r"[^0-9.\-]", "", text)

    if text.count(".") > 1:
        parts = text.split(".")
        text = "".join(parts[:-1]) + "." + parts[-1]

    try:
        return float(text)
    except ValueError:
        return None


def parse_amount_str_to_float(text: str):
    if not text:
        return None

    cleaned = text.strip()
    cleaned = cleaned.replace("\xa0", "")
    cleaned = cleaned.replace("\u202f", "")
    cleaned = cleaned.replace(" ", "")
    cleaned = cleaned.replace("-", ".")
    cleaned = cleaned.replace(",", ".")

    match = re.search(r"(\d+(?:\.\d{1,2})?)", cleaned)
    if not match:
        return None

    try:
        return float(match.group(1))
    except Exception:
        return None


def clean_source_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df = df.dropna(how="all").reset_index(drop=True)

    for col in [
        "ТипДокумента",
        "ВидОперации",
        "Назначение",
        "ДокОснование",
        "Контрагент",
        "Организация",
        "Договор",
        "Комментарий",
        "Ответственный",
        "СтатьяДДС",
        "Валюта",
        "СчетРасчетов",
        "СчетКонтр",
        "БанкКонтр",
        "КодКонтр",
        "Номер1С",
        "НомерБанк",
    ]:
        if col in df.columns:
            df[col] = df[col].apply(normalize_text)

    for col in ["ИНН_Орг", "ИНН_Контр"]:
        if col in df.columns:
            df[col] = df[col].apply(normalize_inn)

    for col in ["КПП_Орг", "КПП_Контр"]:
        if col in df.columns:
            df[col] = df[col].apply(normalize_kpp)

    for col in ["НашСчет", "СчетКонтр"]:
        if col in df.columns:
            df[col] = df[col].apply(normalize_account)

    for col in ["НашБИК", "БИК_Контр"]:
        if col in df.columns:
            df[col] = df[col].apply(normalize_bik)

    if "Сумма" in df.columns:
        df["Сумма"] = df["Сумма"].apply(parse_amount)

    if "ДатаБанк" in df.columns:
        df["ДатаБанк"] = df["ДатаБанк"].apply(normalize_date_for_excel)

    if "Дата1С" in df.columns:
        df["Дата1С"] = df["Дата1С"].apply(normalize_date_for_excel)

    def keep_row(row):
        return any([
            normalize_text(row.get("ТипДокумента", "")),
            normalize_text(row.get("ВидОперации", "")),
            row.get("Сумма", None) is not None and not pd.isna(row.get("Сумма", None)),
        ])

    df = df[df.apply(keep_row, axis=1)].reset_index(drop=True)
    return df


def build_mapping_dict(df_mapping: pd.DataFrame):
    mapping_dict = {}
    duplicates = []

    for _, row in df_mapping.iterrows():
        key = (row["ТипДокументаБП_norm"], row["ВидОперацииБП_norm"])
        if key in mapping_dict:
            duplicates.append(key)
        mapping_dict[key] = {
            "ДокументУТ": row["ДокументУТ"],
            "ОперацияУТ": row["ОперацияУТ"],
            "ГруппаУчета": row["ГруппаУчета"],
            "НаправлениеДенег": row["НаправлениеДенег"],
            "ТребуетПроверки": row["ТребуетПроверки"],
            "КомментарийМаппинга": row["КомментарийМаппинга"],
        }

    if duplicates:
        raise ValueError(
            "В файле маппинга найдены дубли по ключу "
            "(ТипДокументаБП + ВидОперацииБП): "
            f"{duplicates}"
        )

    return mapping_dict


def build_dds_mapping_dict(df_mapping: pd.DataFrame):
    mapping_dict = {}
    duplicates = []

    for _, row in df_mapping.iterrows():
        key = (row["ТипДокументаБП_norm"], row["ВидОперацииБП_norm"])
        if key in mapping_dict:
            duplicates.append(key)
        mapping_dict[key] = {
            "СтатьяДДС_УТ": row["СтатьяДДС_УТ"],
        }

    if duplicates:
        raise ValueError(
            "В файле маппинга статей ДДС найдены дубли по ключу "
            "(ТипДокументаБП + ВидОперацииБП): "
            f"{duplicates}"
        )

    return mapping_dict


def get_default_vat_by_date(doc_date):
    border_date = datetime(2019, 1, 1).date()
    if doc_date and doc_date < border_date:
        return "18%"
    return "20%"


def detect_vat(text: str, doc_date=None):
    source = normalize_text(text)
    source_lower = source.lower()

    if not source:
        return {
            "НДС_ИзТекста": "Без НДС",
            "СуммаНДС_ИзТекста": None,
            "ИсточникНДС": "Нет упоминания НДС в тексте",
        }

    no_vat_patterns = [
        r"\bбез\s+ндс\b",
        r"ндс\s+не\s+облагается",
        r"без\s+налога\s*\(?\s*ндс\s*\)?",
        r"без\s+налога",
        r"не\s+облагается",
    ]

    for pattern in no_vat_patterns:
        if re.search(pattern, source_lower, flags=re.IGNORECASE):
            return {
                "НДС_ИзТекста": "Без НДС",
                "СуммаНДС_ИзТекста": None,
                "ИсточникНДС": "Распознано по тексту",
            }

    explicit_patterns = [
        r"ндс[^\d]{0,15}\(?\s*(20|18|10|7|5|0)\s*%\s*\)?\s*[-:–—]?\s*([\d\s]+(?:[.,-]\d{2})?)",
        r"в\s*т\.?\s*ч\.?\s*ндс[^\d]{0,15}\(?\s*(20|18|10|7|5|0)\s*%\s*\)?\s*[-:–—]?\s*([\d\s]+(?:[.,-]\d{2})?)",
        r"в\s*том\s*числе\s*ндс[^\d]{0,15}\(?\s*(20|18|10|7|5|0)\s*%\s*\)?\s*[-:–—]?\s*([\d\s]+(?:[.,-]\d{2})?)",
    ]

    for pattern in explicit_patterns:
        m = re.search(pattern, source_lower, flags=re.IGNORECASE)
        if m:
            rate = f"{m.group(1)}%"
            amount = parse_amount_str_to_float(m.group(2))
            return {
                "НДС_ИзТекста": rate,
                "СуммаНДС_ИзТекста": amount,
                "ИсточникНДС": "Распознано по ставке и сумме",
            }

    rate_patterns = [
        r"ндс[^\d]{0,15}\(?\s*(20|18|10|7|5|0)\s*%\s*\)?",
        r"в\s*т\.?\s*ч\.?\s*ндс[^\d]{0,15}\(?\s*(20|18|10|7|5|0)\s*%\s*\)?",
        r"в\s*том\s*числе\s*ндс[^\d]{0,15}\(?\s*(20|18|10|7|5|0)\s*%\s*\)?",
    ]

    for pattern in rate_patterns:
        m = re.search(pattern, source_lower, flags=re.IGNORECASE)
        if m:
            rate = f"{m.group(1)}%"
            return {
                "НДС_ИзТекста": rate,
                "СуммаНДС_ИзТекста": None,
                "ИсточникНДС": "Распознано по ставке",
            }

    amount_only_patterns = [
        r"ндс\s*[-:–—]?\s*([\d\s]+(?:[.,-]\d{2})?)",
        r"в\s*т\.?\s*ч\.?\s*ндс\s*[-:–—]?\s*([\d\s]+(?:[.,-]\d{2})?)",
        r"в\s*том\s*числе\s*ндс\s*[-:–—]?\s*([\d\s]+(?:[.,-]\d{2})?)",
    ]

    for pattern in amount_only_patterns:
        m = re.search(pattern, source_lower, flags=re.IGNORECASE)
        if m:
            amount = parse_amount_str_to_float(m.group(1))
            default_rate = get_default_vat_by_date(doc_date)
            return {
                "НДС_ИзТекста": default_rate,
                "СуммаНДС_ИзТекста": amount,
                "ИсточникНДС": "Есть сумма НДС, ставка определена по дате документа",
            }

    if "ндс" in source_lower and "взимается дополнительно" in source_lower:
        default_rate = get_default_vat_by_date(doc_date)
        return {
            "НДС_ИзТекста": default_rate,
            "СуммаНДС_ИзТекста": None,
            "ИсточникНДС": "НДС указан без ставки, ставка определена по дате документа",
        }

    if "ндс" in source_lower:
        default_rate = get_default_vat_by_date(doc_date)
        return {
            "НДС_ИзТекста": default_rate,
            "СуммаНДС_ИзТекста": None,
            "ИсточникНДС": "Есть упоминание НДС, ставка определена по дате документа",
        }

    return {
        "НДС_ИзТекста": "Без НДС",
        "СуммаНДС_ИзТекста": None,
        "ИсточникНДС": "Нет упоминания НДС в тексте",
    }


def apply_mapping(df: pd.DataFrame, mapping_dict: dict, dds_mapping_dict: dict) -> pd.DataFrame:
    df = df.copy()

    new_cols = [
        "ДокументУТ",
        "ОперацияУТ",
        "ГруппаУчета",
        "НаправлениеДенег",
        "ТребуетПроверки",
        "КомментарийМаппинга",
        "НДС_ИзТекста",
        "СуммаНДС_ИзТекста",
        "ИсточникНДС",
        "МаппингНайден",
        "СтатьяДДС_УТ",
    ]

    for col in new_cols:
        if col not in df.columns:
            df[col] = None

    for idx, row in df.iterrows():
        doc_type = normalize_key(row.get("ТипДокумента", ""))
        operation = normalize_key(row.get("ВидОперации", ""))
        key = (doc_type, operation)

        if key in mapping_dict:
            mapped = mapping_dict[key]
            for col, value in mapped.items():
                df.at[idx, col] = value
            df.at[idx, "МаппингНайден"] = "Да"
        else:
            df.at[idx, "ДокументУТ"] = None
            df.at[idx, "ОперацияУТ"] = None
            df.at[idx, "ГруппаУчета"] = None
            df.at[idx, "НаправлениеДенег"] = None
            df.at[idx, "ТребуетПроверки"] = "Да"
            df.at[idx, "КомментарийМаппинга"] = "Не найдено соответствие в mapping_operations_ut.xlsx"
            df.at[idx, "МаппингНайден"] = "Нет"

        if key in dds_mapping_dict:
            dds_mapped = dds_mapping_dict[key]
            df.at[idx, "СтатьяДДС_УТ"] = dds_mapped.get("СтатьяДДС_УТ")
        else:
            df.at[idx, "СтатьяДДС_УТ"] = None

        vat_source_text = " ".join([
            normalize_text(row.get("Назначение", "")),
            normalize_text(row.get("ДокОснование", "")),
        ]).strip()

        doc_date = parse_document_date(row.get("ДатаБанк")) or parse_document_date(row.get("Дата1С"))
        vat_result = detect_vat(vat_source_text, doc_date=doc_date)

        df.at[idx, "НДС_ИзТекста"] = vat_result["НДС_ИзТекста"]
        df.at[idx, "СуммаНДС_ИзТекста"] = vat_result["СуммаНДС_ИзТекста"]
        df.at[idx, "ИсточникНДС"] = vat_result["ИсточникНДС"]

    return df


def build_unmapped_report(df: pd.DataFrame) -> pd.DataFrame:
    report = df[df["МаппингНайден"] == "Нет"][["ТипДокумента", "ВидОперации"]].copy()
    if report.empty:
        return pd.DataFrame(columns=["ТипДокумента", "ВидОперации", "Количество"])
    report["Количество"] = 1
    report = (
        report.groupby(["ТипДокумента", "ВидОперации"], dropna=False)["Количество"]
        .sum()
        .reset_index()
        .sort_values(["ТипДокумента", "ВидОперации"])
    )
    return report


def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    summary = [
        {"Показатель": "Всего строк", "Значение": len(df)},
        {"Показатель": "Маппинг найден", "Значение": int((df["МаппингНайден"] == "Да").sum())},
        {"Показатель": "Маппинг не найден", "Значение": int((df["МаппингНайден"] == "Нет").sum())},
        {"Показатель": "Требует проверки = Да", "Значение": int((df["ТребуетПроверки"] == "Да").sum())},
        {"Показатель": "Статья ДДС УТ заполнена", "Значение": int(df["СтатьяДДС_УТ"].fillna("").astype(str).str.strip().ne("").sum())},
        {"Показатель": "Статья ДДС УТ пустая", "Значение": int(df["СтатьяДДС_УТ"].fillna("").astype(str).str.strip().eq("").sum())},
        {"Показатель": "НДС = Без НДС", "Значение": int((df["НДС_ИзТекста"] == "Без НДС").sum())},
        {"Показатель": "НДС = 20%", "Значение": int((df["НДС_ИзТекста"] == "20%").sum())},
        {"Показатель": "НДС = 18%", "Значение": int((df["НДС_ИзТекста"] == "18%").sum())},
        {"Показатель": "НДС = 10%", "Значение": int((df["НДС_ИзТекста"] == "10%").sum())},
        {"Показатель": "НДС = 7%", "Значение": int((df["НДС_ИзТекста"] == "7%").sum())},
        {"Показатель": "НДС = 5%", "Значение": int((df["НДС_ИзТекста"] == "5%").sum())},
    ]
    return pd.DataFrame(summary)


def autosize_worksheet_columns(worksheet, df: pd.DataFrame):
    from openpyxl.utils import get_column_letter

    for idx, column_name in enumerate(df.columns, start=1):
        max_len = len(str(column_name))
        series = df[column_name].astype(str).fillna("")
        if not series.empty:
            max_len = max(max_len, series.map(len).max())
        worksheet.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 60)


def apply_excel_formats(writer, df_data: pd.DataFrame):
    ws = writer.sheets["Данные"]

    date_columns = {"Дата1С", "ДатаБанк"}
    amount_columns = {"Сумма", "СуммаНДС_ИзТекста"}

    col_index = {name: idx + 1 for idx, name in enumerate(df_data.columns)}

    for col_name in date_columns:
        if col_name in col_index:
            col_num = col_index[col_name]
            for row in range(2, len(df_data) + 2):
                ws.cell(row=row, column=col_num).number_format = "DD.MM.YYYY"

    for col_name in amount_columns:
        if col_name in col_index:
            col_num = col_index[col_name]
            for row in range(2, len(df_data) + 2):
                ws.cell(row=row, column=col_num).number_format = "#,##0.00"

    autosize_worksheet_columns(ws, df_data)

    if "НераспознанныеОперации" in writer.sheets:
        autosize_worksheet_columns(
            writer.sheets["НераспознанныеОперации"],
            build_unmapped_report(df_data)
        )

    if "Сводка" in writer.sheets:
        autosize_worksheet_columns(
            writer.sheets["Сводка"],
            build_summary(df_data)
        )


def build_output_filename(year: int, min_date: date, max_date: date) -> str:
    start_date = date(year, 1, 1)

    if max_date.year == year:
        end_date = max_date
    else:
        end_date = date(year, 12, 31)

    return f"Выгрузка_Банка_{start_date:%Y%m%d}_{end_date:%Y%m%d}_prepared.xlsx"


def save_single_result(output_path: Path, df_result: pd.DataFrame):
    unmapped_report = build_unmapped_report(df_result)
    summary = build_summary(df_result)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_result.to_excel(writer, sheet_name="Данные", index=False)
        unmapped_report.to_excel(writer, sheet_name="НераспознанныеОперации", index=False)
        summary.to_excel(writer, sheet_name="Сводка", index=False)
        apply_excel_formats(writer, df_result)


def save_results_by_year(input_path: Path, df_result: pd.DataFrame):
    if "ДатаБанк" not in df_result.columns:
        raise ValueError("В результирующем наборе отсутствует колонка 'ДатаБанк'.")

    df_valid = df_result[df_result["ДатаБанк"].notna()].copy()
    df_invalid = df_result[df_result["ДатаБанк"].isna()].copy()

    if df_valid.empty and df_invalid.empty:
        raise ValueError("После обработки не осталось данных для сохранения.")

    saved_files = []

    if not df_valid.empty:
        df_valid["Год"] = df_valid["ДатаБанк"].dt.year

        for year in sorted(df_valid["Год"].dropna().unique()):
            year = int(year)
            year_df = df_valid[df_valid["Год"] == year].copy()
            year_df = year_df.drop(columns=["Год"])

            min_date = year_df["ДатаБанк"].min().date()
            max_date = year_df["ДатаБанк"].max().date()

            output_name = build_output_filename(year, min_date, max_date)
            output_path = input_path.with_name(output_name)

            save_single_result(output_path, year_df)
            saved_files.append(output_path)

    if not df_invalid.empty:
        output_name = f"{input_path.stem}_without_valid_bank_date_prepared.xlsx"
        output_path = input_path.with_name(output_name)
        save_single_result(output_path, df_invalid)
        saved_files.append(output_path)

    return saved_files


def main():
    try:
        input_path = ask_file_path()
        df_source = load_excel_first_sheet(input_path)
        validate_bank_file(df_source)

        df_source = clean_source_dataframe(df_source)

        df_mapping = load_mapping(MAPPING_FILE)
        mapping_dict = build_mapping_dict(df_mapping)

        df_dds_mapping = load_dds_mapping(DDS_MAPPING_FILE)
        dds_mapping_dict = build_dds_mapping_dict(df_dds_mapping)

        df_result = apply_mapping(df_source, mapping_dict, dds_mapping_dict)
        saved_files = save_results_by_year(input_path, df_result)

        files_text = "\n".join(str(path) for path in saved_files)

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)

        messagebox.showinfo(
            "Обработка завершена",
            f"Файлы успешно обработаны.\n\nРезультаты сохранены:\n{files_text}"
        )

        root.destroy()

        print("Результаты сохранены:")
        for path in saved_files:
            print(path)

    except Exception as e:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        messagebox.showerror("Ошибка", str(e))
        root.destroy()

        print("\nОШИБКА:")
        print(str(e))
        sys.exit(1)


if __name__ == "__main__":
    main()