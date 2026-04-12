"""
Скрипт выполняет два действия:

1. Лист "Клиенты который нет в УТ" / tbl_UT_New
   Находит клиентов из tbl_ACC (БП), которых нет в tbl_UT (УТ), и
   записывает их на отдельный лист.

2. Лист "UT_bank" / tbl_UT_Bank
   Берёт всех клиентов из tbl_UT + tbl_UT_New, оставляет только тех,
   у кого пустые поля банка (БИКБанка, НаименованиеБанка, НомерРасчетногоСчета).
   Для каждого такого клиента пытается найти банковские данные в tbl_ACC по ИНН.
   Если данных нет ни в одной таблице — ставит "данных банка нет".

Сравнение клиентов ведётся по колонке ИНН.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

FILE = r"D:\Programming\1С_Создаём контрагентов и договора\ФайлДороботкиКлеинтов_в_УТ.xlsx"

SHEET_NEW      = "Клиенты который нет в УТ"
TABLE_NEW      = "tbl_UT_New"

SHEET_BANK     = "UT_bank"
TABLE_BANK     = "tbl_UT_Bank"

# Колонки с банковскими реквизитами (должны совпадать с заголовками)
BANK_COLS = ["НомерРасчетногоСчета", "БИКБанка", "НаименованиеБанка"]
NO_BANK_MSG = "данных банка нет"


def read_table(wb: openpyxl.Workbook, table_name: str) -> tuple[list, list[list]]:
    """Возвращает (заголовки, строки данных) для именованной таблицы."""
    for ws in wb.worksheets:
        if table_name in ws.tables:
            tbl = ws.tables[table_name]
            cells = list(ws[tbl.ref])
            headers = [cell.value for cell in cells[0]]
            rows = [[cell.value for cell in row] for row in cells[1:]]
            return headers, rows
    raise ValueError(f"Таблица '{table_name}' не найдена в файле")


def is_empty(val) -> bool:
    return val is None or str(val).strip() == ""


def write_sheet(wb: openpyxl.Workbook, sheet_name: str, table_name: str,
                headers: list, rows: list[list]) -> None:
    """Удаляет старый лист (если есть), создаёт новый, пишет данные и таблицу."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)
    ws.append(headers)
    for row in rows:
        ws.append(row)

    n_rows = len(rows) + 1  # +1 строка заголовка
    last_col = get_column_letter(len(headers))
    tbl = Table(displayName=table_name, ref=f"A1:{last_col}{n_rows}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)

    # Автоширина колонок
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header)) if header else 10
        for row in rows:
            val = row[col_idx - 1]
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def main():
    wb = openpyxl.load_workbook(FILE)

    headers_ut,  rows_ut  = read_table(wb, "tbl_UT")
    headers_acc, rows_acc = read_table(wb, "tbl_ACC")

    idx_inn_ut  = headers_ut.index("ИНН")
    idx_inn_acc = headers_acc.index("ИНН")

    # ─── Шаг 1: клиенты из БП, которых нет в УТ ───────────────────────────
    ut_inns = {str(r[idx_inn_ut]).strip() for r in rows_ut if r[idx_inn_ut]}

    missing_rows = [
        r for r in rows_acc
        if r[idx_inn_acc] and str(r[idx_inn_acc]).strip() not in ut_inns
    ]

    print(f"Клиентов в tbl_UT:                {len(rows_ut)}")
    print(f"Клиентов в tbl_ACC:               {len(rows_acc)}")
    print(f"Есть в БП, нет в УТ:              {len(missing_rows)}")

    write_sheet(wb, SHEET_NEW, TABLE_NEW, headers_ut, missing_rows)
    print(f'Лист "{SHEET_NEW}" -> таблица {TABLE_NEW} записана.')

    # ─── Шаг 2: клиенты без банковских реквизитов ──────────────────────────
    # Индексы банковских колонок в структуре tbl_UT
    bank_idxs = [headers_ut.index(col) for col in BANK_COLS]

    # Справочник банковских данных из tbl_ACC по ИНН
    acc_bank_by_inn: dict[str, list] = {}
    for row in rows_acc:
        inn = str(row[idx_inn_acc]).strip() if row[idx_inn_acc] else None
        if inn:
            acc_bank_by_inn[inn] = row  # берём первую найденную запись

    # Все клиенты УТ + новые (из missing_rows, которые имеют ту же структуру)
    all_ut_rows = rows_ut + missing_rows

    bank_rows = []
    for row in all_ut_rows:
        # Пропускаем клиентов у которых все банковские поля уже заполнены
        if not any(is_empty(row[i]) for i in bank_idxs):
            continue

        filled = list(row)  # копия строки
        inn = str(row[idx_inn_ut]).strip() if row[idx_inn_ut] else None
        acc_row = acc_bank_by_inn.get(inn) if inn else None

        for i, col in zip(bank_idxs, BANK_COLS):
            if is_empty(filled[i]):
                idx_in_acc = headers_acc.index(col)
                src_val = acc_row[idx_in_acc] if acc_row else None
                filled[i] = src_val if not is_empty(src_val) else NO_BANK_MSG

        bank_rows.append(filled)

    print(f"Клиентов без банковских данных:   {len(bank_rows)}")

    write_sheet(wb, SHEET_BANK, TABLE_BANK, headers_ut, bank_rows)
    print(f'Лист "{SHEET_BANK}" -> таблица {TABLE_BANK} записана.')

    wb.save(FILE)
    print("\nФайл сохранён.")

    # Итоговый вывод по банковскому листу
    if bank_rows:
        name_idx = headers_ut.index("НаименованиеКонтрагента")
        print(f"\nКлиенты на листе {SHEET_BANK}:")
        for r in bank_rows:
            bik = r[bank_idxs[1]]
            acc = r[bank_idxs[0]]
            print(f"  {r[name_idx]} | БИК: {bik} | Счёт: {acc}")


if __name__ == "__main__":
    main()
