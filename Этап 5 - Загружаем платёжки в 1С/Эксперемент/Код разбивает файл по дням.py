from openpyxl import load_workbook, Workbook
from copy import copy
import os

file_name = "Эксперемент 2018.xlsx"
wb = load_workbook(file_name)
ws = wb.active

folder_path = os.path.dirname(os.path.abspath(file_name))

# Найти колонку
header = [cell.value for cell in ws[1]]
date_col_index = header.index("Дата1С") + 1

data_by_date = {}

for row in ws.iter_rows(min_row=2):
    cell = row[date_col_index - 1]
    date_value = cell.value

    if date_value is None:
        continue

    date_key = date_value.date() if hasattr(date_value, "date") else date_value

    data_by_date.setdefault(date_key, []).append(row)

header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]

# Функция безопасного копирования стиля
def copy_cell(source_cell, target_cell):
    target_cell.value = source_cell.value

    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

# Создание файлов
for date, rows in data_by_date.items():
    new_wb = Workbook()
    new_ws = new_wb.active

    # Заголовок
    for col_idx, cell in enumerate(header_row, start=1):
        new_cell = new_ws.cell(row=1, column=col_idx)
        copy_cell(cell, new_cell)

    # Данные
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, cell in enumerate(row, start=1):
            new_cell = new_ws.cell(row=row_idx, column=col_idx)
            copy_cell(cell, new_cell)

    # Имя файла
    month = str(date.month).zfill(2)
    day = str(date.day).zfill(2)

    new_file_name = f"Эксперемент_2018 {month}_{day}.xlsx"
    new_file_path = os.path.join(folder_path, new_file_name)

    new_wb.save(new_file_path)

print("Готово без падений и с сохранением форматов.")