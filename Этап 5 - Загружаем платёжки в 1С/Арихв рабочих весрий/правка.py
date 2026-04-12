import os
from openpyxl import load_workbook, Workbook
from tkinter import Tk, filedialog
from tqdm import tqdm

def is_empty(value):
    return value is None or value == ""

# --- file picker ---
root = Tk()
root.withdraw()
file_path = filedialog.askopenfilename(
    title="Select Excel file",
    filetypes=[("Excel files", "*.xlsx")]
)

if not file_path:
    print("No file selected")
    exit()

print("Loading workbook...")

wb = load_workbook(file_path)
ws = wb.active

max_row = ws.max_row
max_col = ws.max_column

# Columns
COL_INN = 15  # O
COL_KPP = 16  # P
COL_X = 24    # X
COL_Y = 25    # Y
COL_Z = 26    # Z

client_values = {}
client_has_any_xyz = {}

print("Step 1/3: Collecting client data...")

# --- pass 1 ---
for row in tqdm(ws.iter_rows(min_row=2, max_row=max_row), total=max_row - 1):
    inn = row[COL_INN - 1].value
    kpp = row[COL_KPP - 1].value
    key = (inn, kpp)

    x = row[COL_X - 1].value
    y = row[COL_Y - 1].value
    z = row[COL_Z - 1].value

    if key not in client_has_any_xyz:
        client_has_any_xyz[key] = False

    if not is_empty(x) or not is_empty(y) or not is_empty(z):
        client_has_any_xyz[key] = True

    if key not in client_values:
        if not is_empty(x) or not is_empty(y) or not is_empty(z):
            client_values[key] = (x, y, z)

print("Step 2/3: Filling missing values...")

# --- pass 2 ---
for row in tqdm(ws.iter_rows(min_row=2, max_row=max_row), total=max_row - 1):
    inn = row[COL_INN - 1].value
    kpp = row[COL_KPP - 1].value
    key = (inn, kpp)

    if key in client_values:
        saved_x, saved_y, saved_z = client_values[key]

        if is_empty(row[COL_X - 1].value):
            row[COL_X - 1].value = saved_x
        if is_empty(row[COL_Y - 1].value):
            row[COL_Y - 1].value = saved_y
        if is_empty(row[COL_Z - 1].value):
            row[COL_Z - 1].value = saved_z

print("Saving main file...")

dir_name = os.path.dirname(file_path)
base_name = os.path.splitext(os.path.basename(file_path))[0]

output_main = os.path.join(dir_name, base_name + "_filled.xlsx")
wb.save(output_main)

# =========================
# FAST STEP 3 (FIXED)
# =========================

print("Step 3/3: Creating file with empty clients (fast)...")

wb_empty = Workbook()
ws_empty = wb_empty.active

# --- copy header ---
for col in range(1, max_col + 1):
    ws_empty.cell(row=1, column=col).value = ws.cell(row=1, column=col).value

new_row = 2

# --- copy only needed rows ---
for row in tqdm(ws.iter_rows(min_row=2, max_row=max_row), total=max_row - 1):
    inn = row[COL_INN - 1].value
    kpp = row[COL_KPP - 1].value
    key = (inn, kpp)

    if not client_has_any_xyz.get(key, False):
        for col_idx, cell in enumerate(row, start=1):
            ws_empty.cell(row=new_row, column=col_idx).value = cell.value
        new_row += 1

output_empty = os.path.join(dir_name, base_name + "_no_xyz.xlsx")
wb_empty.save(output_empty)

print("\nDone!")
print("Main file:", output_main)
print("Empty clients file:", output_empty)